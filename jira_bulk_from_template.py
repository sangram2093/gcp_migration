#!/usr/bin/env python3
"""
Bulk Jira creator for feed/scenario migration using:
1) template Excel (jira_board_tasks.xlsx style)
2) metadata JSON (feeds/scenarios/project settings)

Features:
- dry-run mode
- retries with backoff for Jira API calls
- checkpoint/resume via state file
- robust text sanitization for quotes/newlines from Excel
- acceptance criteria set as separate custom field
- story->feature linking with link-type name resolution
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
import base64
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import requests
from openpyxl import load_workbook
from requests.auth import HTTPBasicAuth


REQUIRED_COLUMNS = [
    "Feature",
    "Feature Description",
    "Feature Acceptance Criteria",
    "Story",
    "Story Description",
    "Sub-Task",
    "Sub-Task Description",
    "Sub-Task Acceptance Criteria",
]

FORWARD_FILL_COLUMNS = [
    "Feature",
    "Feature Description",
    "Feature Acceptance Criteria",
    "Story",
    "Story Description",
]


def sanitize_text(value: Any, multiline: bool = True) -> str:
    if value is None:
        return ""
    text = str(value)
    text = (
        text.replace("\u2018", "'")
        .replace("\u2019", "'")
        .replace("\u201c", '"')
        .replace("\u201d", '"')
        .replace("\u00a0", " ")
    )
    text = text.replace("\\r\\n", "\n").replace("\\n", "\n").replace("\\t", "\t")
    text = text.replace('\\"', '"').replace("\\'", "'")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = re.sub(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]", "", text)
    text = text.strip()
    # Remove outer matching quotes repeatedly
    while len(text) >= 2 and (
        (text[0] == text[-1] and text[0] in ['"', "'", "`"])
        or (text.startswith("\u201c") and text.endswith("\u201d"))
    ):
        text = text[1:-1].strip()

    if multiline:
        lines = [ln.rstrip() for ln in text.split("\n")]
        text = "\n".join(lines)
        text = re.sub(r"\n{3,}", "\n\n", text).strip()
    else:
        text = re.sub(r"\s+", " ", text).strip()

    return text


def sanitize_key(value: Any) -> str:
    text = sanitize_text(value, multiline=False)
    text = re.sub(r"[.,;:]+$", "", text)
    text = re.sub(r"\s+", "", text)
    return text


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Bulk Jira creation from template + metadata")
    parser.add_argument("--template", required=True, help="Path to template Excel")
    parser.add_argument("--metadata", required=True, help="Path to metadata JSON")
    parser.add_argument("--sheet", default=None, help="Excel sheet name (default: active)")
    parser.add_argument("--feed-rows", type=int, default=10, help="Number of feed template rows")
    parser.add_argument("--scenario-rows", type=int, default=15, help="Number of scenario template rows")
    parser.add_argument("--jira-url", default=os.getenv("JIRA_URL", ""), help="Jira base URL (e.g. https://company.atlassian.net)")
    parser.add_argument("--jira-email", default=os.getenv("JIRA_EMAIL", ""), help="Jira user email")
    parser.add_argument("--jira-token", default=os.getenv("JIRA_API_TOKEN", ""), help="Jira API token")
    parser.add_argument(
        "--jira-auth-mode",
        default=os.getenv("JIRA_AUTH_MODE", ""),
        choices=["", "basic", "bearer", "auto"],
        help="Jira auth mode: bearer (default, same as built-in tools), basic, or auto",
    )
    parser.add_argument(
        "--jira-api-version",
        default=os.getenv("JIRA_API_VERSION", ""),
        choices=["", "2", "3"],
        help="Jira REST API version (default: 2, same as built-in tools)",
    )
    parser.add_argument("--state-file", default=".jira_bulk_state.json", help="Checkpoint file path")
    parser.add_argument("--dry-run", action="store_true", help="Print plan, do not create issues")
    parser.add_argument("--auth-debug", action="store_true", help="Print safe auth diagnostics (no token value)")
    parser.add_argument("--max-retries", type=int, default=5, help="API retry count")
    parser.add_argument("--retry-backoff-seconds", type=float, default=1.5, help="Retry backoff base")
    return parser.parse_args()


def load_metadata(path: str) -> Dict[str, Any]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    required = ["projectKey", "epicKeyForFeatures", "surveillanceName"]
    for key in required:
        if not data.get(key):
            raise ValueError(f"metadata missing required field: {key}")
    data.setdefault("feeds", [])
    data.setdefault("scenarios", [])
    data.setdefault("labels", [])
    data.setdefault("linkType", "Relates")
    return data


def read_template_rows(path: str, sheet_name: Optional[str]) -> List[Dict[str, str]]:
    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    header_row = None
    header_map: Dict[str, int] = {}
    for row_idx in range(1, ws.max_row + 1):
        values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        normalized = [sanitize_text(v, multiline=False) for v in values]
        if not any(normalized):
            continue
        candidate = {name: idx for idx, name in enumerate(normalized, start=1) if name}
        if all(col in candidate for col in REQUIRED_COLUMNS):
            header_row = row_idx
            header_map = {col: candidate[col] for col in REQUIRED_COLUMNS}
            break
    if header_row is None:
        raise ValueError(f"Could not find required columns in template: {REQUIRED_COLUMNS}")

    rows: List[Dict[str, str]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        row_data: Dict[str, str] = {}
        any_value = False
        for col in REQUIRED_COLUMNS:
            raw = ws.cell(row=row_idx, column=header_map[col]).value
            val = sanitize_text(raw, multiline=True)
            row_data[col] = val
            if val:
                any_value = True
        if any_value:
            rows.append(row_data)
    if not rows:
        raise ValueError("Template has no data rows after header")

    # Forward fill merged-like blank fields
    last_seen: Dict[str, str] = {k: "" for k in FORWARD_FILL_COLUMNS}
    for row in rows:
        for col in FORWARD_FILL_COLUMNS:
            if row[col]:
                last_seen[col] = row[col]
            else:
                row[col] = last_seen[col]
    return rows


@dataclass
class TemplateSection:
    feature: str
    feature_description: str
    feature_ac: str
    story: str
    story_description: str
    subtasks: List[Tuple[str, str, str]]


def section_from_rows(rows: List[Dict[str, str]]) -> TemplateSection:
    first = rows[0]
    subtasks: List[Tuple[str, str, str]] = []
    for row in rows:
        if row["Sub-Task"]:
            subtasks.append(
                (
                    row["Sub-Task"],
                    row["Sub-Task Description"],
                    row["Sub-Task Acceptance Criteria"],
                )
            )
    return TemplateSection(
        feature=first["Feature"],
        feature_description=first["Feature Description"],
        feature_ac=first["Feature Acceptance Criteria"],
        story=first["Story"],
        story_description=first["Story Description"],
        subtasks=subtasks,
    )


def apply_placeholders(text: str, surveillance_name: str, scenario_name: str = "", feed_name: str = "") -> str:
    out = text
    out = out.replace("SURVEILLANCE_NAME", surveillance_name)
    out = out.replace("SCENARIO_NAME", scenario_name)
    out = out.replace("FEED_NAME", feed_name)
    return sanitize_text(out, multiline=True)


def _first_non_empty(values: List[str]) -> str:
    for value in values:
        if sanitize_text(value, multiline=False):
            return sanitize_text(value, multiline=False)
    return ""


def resolve_jira_credentials(args: argparse.Namespace, metadata: Dict[str, Any]) -> Tuple[str, str, str, str, str]:
    jira_meta = metadata.get("jira", {}) if isinstance(metadata.get("jira"), dict) else {}
    jira_url = _first_non_empty(
        [
            args.jira_url,
            os.getenv("JIRA_URL", ""),
            os.getenv("JIRA_BASE_URL", ""),
            jira_meta.get("url", ""),
            jira_meta.get("baseUrl", ""),
            jira_meta.get("domain", ""),
        ]
    )
    jira_email = _first_non_empty(
        [
            args.jira_email,
            os.getenv("JIRA_EMAIL", ""),
            os.getenv("JIRA_AUTH_EMAIL", ""),
            jira_meta.get("email", ""),
            jira_meta.get("authEmail", ""),
        ]
    )
    jira_token = _first_non_empty(
        [
            args.jira_token,
            os.getenv("JIRA_API_TOKEN", ""),
            os.getenv("JIRA_TOKEN", ""),
            os.getenv("JIRA_PAT", ""),
            jira_meta.get("token", ""),
            jira_meta.get("apiToken", ""),
            jira_meta.get("pat", ""),
        ]
    )
    jira_auth_mode = _first_non_empty(
        [
            args.jira_auth_mode,
            os.getenv("JIRA_AUTH_MODE", ""),
            jira_meta.get("authMode", ""),
            "bearer",
        ]
    ).lower()
    if jira_auth_mode not in ("basic", "bearer", "auto"):
        jira_auth_mode = "bearer"
    jira_api_version = _first_non_empty(
        [
            args.jira_api_version,
            os.getenv("JIRA_API_VERSION", ""),
            str(jira_meta.get("apiVersion", "")),
            "2",
        ]
    )
    if jira_api_version not in ("2", "3"):
        jira_api_version = "2"
    return jira_url, jira_email, jira_token, jira_auth_mode, jira_api_version


class JiraClient:
    def __init__(
        self,
        base_url: str,
        email: str,
        token: str,
        auth_mode: str,
        api_version: str,
        max_retries: int,
        backoff: float,
    ) -> None:
        normalized_base = sanitize_text(base_url, multiline=False)
        if not re.match(r"^https?://", normalized_base, flags=re.IGNORECASE):
            normalized_base = f"https://{normalized_base}"
        # Allow users to pass either site root or rest-api root
        normalized_base = re.sub(
            r"/rest/api/[23]/?$",
            "",
            normalized_base,
            flags=re.IGNORECASE,
        )
        self.base_url = normalized_base.rstrip("/")
        self.auth_mode = sanitize_text(auth_mode, multiline=False).lower() or "bearer"
        if self.auth_mode not in ("basic", "bearer", "auto"):
            self.auth_mode = "bearer"
        self.api_version = sanitize_text(api_version, multiline=False) or "2"
        if self.api_version not in ("2", "3"):
            self.api_version = "2"
        self.api_prefix = f"/rest/api/{self.api_version}"
        self.email = sanitize_text(email, multiline=False)
        self.token = sanitize_text(token, multiline=False)
        self.auth: Optional[HTTPBasicAuth] = None
        self.auth_header: Optional[str] = None
        if self.auth_mode == "basic":
            self.auth = HTTPBasicAuth(self.email, self.token)
            basic_payload = f"{self.email}:{self.token}".encode("utf-8")
            self.auth_header = f"Basic {base64.b64encode(basic_payload).decode('ascii')}"
        elif self.auth_mode == "bearer":
            self.auth_header = f"Bearer {self.token}"
        else:  # auto
            self.auth_header = f"Bearer {self.token}"
        self.max_retries = max_retries
        self.backoff = backoff
        self.session = requests.Session()
        self.field_cache: Dict[str, str] = {}
        self.link_type_cache: Optional[List[Dict[str, str]]] = None
        self.issue_type_cache: Dict[str, Dict[str, Any]] = {}

    def _request(self, method: str, path: str, **kwargs: Any) -> Any:
        url = f"{self.base_url}{path}"
        headers = kwargs.pop("headers", {})
        headers.setdefault("Accept", "application/json")
        if self.auth_header:
            headers.setdefault("Authorization", self.auth_header)
        if "json" in kwargs:
            headers.setdefault("Content-Type", "application/json")
        last_error: Optional[str] = None
        for attempt in range(1, self.max_retries + 1):
            try:
                resp = self.session.request(
                    method=method,
                    url=url,
                    auth=self.auth,
                    headers=headers,
                    timeout=45,
                    **kwargs,
                )
                if resp.status_code in (429, 500, 502, 503, 504):
                    last_error = f"{resp.status_code} {resp.text}"
                    time.sleep(self.backoff * attempt)
                    continue
                if resp.status_code == 401 and self.auth_mode == "auto":
                    # Auto-fallback: try basic if bearer failed and email is provided.
                    if headers.get("Authorization", "").startswith("Bearer ") and self.email:
                        basic_payload = f"{self.email}:{self.token}".encode("utf-8")
                        headers["Authorization"] = f"Basic {base64.b64encode(basic_payload).decode('ascii')}"
                        resp = self.session.request(
                            method=method,
                            url=url,
                            auth=None,
                            headers=headers,
                            timeout=45,
                            **kwargs,
                        )
                        if resp.status_code < 400:
                            if resp.status_code == 204 or not resp.text:
                                return {}
                            return resp.json()
                if resp.status_code >= 400:
                    raise RuntimeError(f"Jira API error {resp.status_code} {path}: {resp.text}")
                if resp.status_code == 204 or not resp.text:
                    return {}
                return resp.json()
            except Exception as exc:  # pylint: disable=broad-exception-caught
                last_error = str(exc)
                if attempt < self.max_retries:
                    time.sleep(self.backoff * attempt)
                    continue
                raise RuntimeError(f"Jira API request failed after retries: {method} {path} :: {last_error}") from exc
        raise RuntimeError(f"Jira API request failed: {method} {path} :: {last_error}")

    def get_field_id(self, field_name: str) -> str:
        if field_name in self.field_cache:
            return self.field_cache[field_name]
        fields = self._request("GET", f"{self.api_prefix}/field")
        for field in fields:
            if sanitize_text(field.get("name", ""), multiline=False).lower() == field_name.lower():
                field_id = field["id"]
                self.field_cache[field_name] = field_id
                return field_id
        raise RuntimeError(f"Jira field not found: {field_name}")

    def _load_project_issue_types(self, project_key: str) -> List[Dict[str, Any]]:
        cache_key = sanitize_key(project_key)
        if cache_key in self.issue_type_cache:
            return self.issue_type_cache[cache_key]["items"]
        data = self._request(
            "GET",
            f"{self.api_prefix}/issue/createmeta?projectKeys={cache_key}&expand=projects.issuetypes",
        )
        projects = data.get("projects", []) if isinstance(data, dict) else []
        if not projects:
            raise RuntimeError(f"No Jira create metadata for project {cache_key}")
        issue_types = projects[0].get("issuetypes", []) or []
        self.issue_type_cache[cache_key] = {"items": issue_types}
        return issue_types

    def get_subtask_type_id(self, project_key: str) -> str:
        cache_key = sanitize_key(project_key)
        data = self._request(
            "GET",
            f"{self.api_prefix}/issue/createmeta?projectKeys={cache_key}&expand=projects.issuetypes.fields",
        )
        projects = data.get("projects", []) if isinstance(data, dict) else []
        if not projects:
            raise RuntimeError(f"No Jira create metadata for project {cache_key}")
        issue_types = projects[0].get("issuetypes", []) or []
        for it in issue_types:
            if bool(it.get("subtask", False)):
                return str(it["id"])
        available = [str(it.get("name", "")) for it in issue_types]
        raise RuntimeError(
            f"No sub-task issue type found for project {cache_key}. Available issue types: {available}"
        )

    @staticmethod
    def _normalize_issue_type_name(name: str) -> str:
        return re.sub(r"[^a-z0-9]+", "", sanitize_text(name, multiline=False).lower())

    def resolve_issue_type_id(self, project_key: str, preferred: str, is_subtask: bool = False) -> str:
        issue_types = self._load_project_issue_types(project_key)
        preferred_norm = self._normalize_issue_type_name(preferred)

        aliases = [preferred_norm]
        if preferred_norm in ("newfeature", "feature"):
            aliases.extend(["newfeature", "feature"])
        elif preferred_norm in ("story",):
            aliases.extend(["story", "userstory"])
        elif preferred_norm in ("subtask", "sub-task", "sub task"):
            aliases.extend(["subtask", "sub-task", "subtaskissue"])

        alias_set = set(aliases)

        # exact/alias match first (with subtask flag awareness)
        for it in issue_types:
            name_norm = self._normalize_issue_type_name(it.get("name", ""))
            if name_norm in alias_set and bool(it.get("subtask", False)) == is_subtask:
                return str(it["id"])

        # relaxed match ignoring subtask if no exact found
        for it in issue_types:
            name_norm = self._normalize_issue_type_name(it.get("name", ""))
            if name_norm in alias_set:
                return str(it["id"])

        # fallback by subtask flag only
        for it in issue_types:
            if bool(it.get("subtask", False)) == is_subtask:
                return str(it["id"])

        available = [str(it.get("name", "")) for it in issue_types]
        raise RuntimeError(
            f"Could not resolve issue type '{preferred}' for project {project_key}. Available: {available}"
        )

    def create_issue(
        self,
        project_key: str,
        issue_type: str,
        summary: str,
        description: str,
        epic_key: str = "",
        parent_key: str = "",
        labels: Optional[List[str]] = None,
    ) -> str:
        normalized_project_key = sanitize_key(project_key)
        issue_type_id = ""
        if parent_key:
            # Match built-in Jira tool behavior: always select subtask type by `subtask=true`.
            issue_type_id = self.get_subtask_type_id(normalized_project_key)
        else:
            issue_type_id = self.resolve_issue_type_id(
                normalized_project_key,
                issue_type,
                is_subtask=False,
            )
        fields: Dict[str, Any] = {
            "project": {"key": normalized_project_key},
            "issuetype": {"id": issue_type_id},
            "summary": sanitize_text(summary, multiline=False),
            "description": sanitize_text(description, multiline=True),
        }
        if labels:
            fields["labels"] = [sanitize_text(x, multiline=False) for x in labels if sanitize_text(x, multiline=False)]
        if parent_key:
            fields["parent"] = {"key": sanitize_key(parent_key)}
        if epic_key:
            epic_field = self.get_field_id("Epic Link")
            fields[epic_field] = sanitize_key(epic_key)
        data = self._request("POST", f"{self.api_prefix}/issue", json={"fields": fields})
        key = sanitize_text(data.get("key", ""), multiline=False)
        if not key:
            raise RuntimeError("Jira create issue response did not contain issue key")
        return key

    def set_acceptance_criteria(self, issue_key: str, criteria: str) -> None:
        field_id = self.get_field_id("Acceptance criteria")
        variants = []
        base = ensure_bullets(sanitize_text(criteria, multiline=True))
        variants.append(base)
        variants.append(ensure_bullets(sanitize_text(criteria, multiline=False)))
        variants.append(base.replace('"', "").replace("'", ""))
        seen = set()
        for value in variants:
            value = value.strip()
            if not value or value in seen:
                continue
            seen.add(value)
            try:
                self._request(
                    "PUT",
                    f"{self.api_prefix}/issue/{sanitize_key(issue_key)}",
                    json={"fields": {field_id: value}},
                )
                return
            except Exception:
                continue
        raise RuntimeError(f"Failed to set acceptance criteria for {issue_key}")

    def _load_link_types(self) -> List[Dict[str, str]]:
        if self.link_type_cache is None:
            data = self._request("GET", f"{self.api_prefix}/issueLinkType")
            self.link_type_cache = data.get("issueLinkTypes", []) or []
        return self.link_type_cache

    def resolve_link_type_name(self, preferred: str) -> str:
        preferred_norm = sanitize_text(preferred, multiline=False).lower()
        types = self._load_link_types()
        for entry in types:
            name = sanitize_text(entry.get("name", ""), multiline=False)
            inward = sanitize_text(entry.get("inward", ""), multiline=False)
            outward = sanitize_text(entry.get("outward", ""), multiline=False)
            if preferred_norm in {name.lower(), inward.lower(), outward.lower()}:
                return name
        # fallback if caller already passed a valid name
        return sanitize_text(preferred, multiline=False)

    def link_issues(self, inward_key: str, outward_key: str, link_type: str) -> None:
        resolved_name = self.resolve_link_type_name(link_type)
        candidates = [resolved_name]
        if resolved_name.lower() != sanitize_text(link_type, multiline=False).lower():
            candidates.append(sanitize_text(link_type, multiline=False))
        last_error: Optional[Exception] = None
        for candidate in candidates:
            try:
                self._request(
                    "POST",
                    f"{self.api_prefix}/issueLink",
                    json={
                        "type": {"name": candidate},
                        "inwardIssue": {"key": sanitize_key(inward_key)},
                        "outwardIssue": {"key": sanitize_key(outward_key)},
                    },
                )
                return
            except Exception as exc:  # pylint: disable=broad-exception-caught
                last_error = exc
                continue
        raise RuntimeError(f"Failed to link {inward_key} -> {outward_key}: {last_error}")


class StateStore:
    def __init__(self, path: str) -> None:
        self.path = Path(path)
        self.data: Dict[str, Any] = {"issues": {}, "links": {}}
        if self.path.exists():
            self.data = json.loads(self.path.read_text(encoding="utf-8"))

    def save(self) -> None:
        self.path.write_text(json.dumps(self.data, indent=2), encoding="utf-8")

    def get_issue(self, logical_id: str) -> Optional[str]:
        return self.data["issues"].get(logical_id)

    def set_issue(self, logical_id: str, issue_key: str) -> None:
        self.data["issues"][logical_id] = issue_key
        self.save()

    def has_link(self, logical_id: str) -> bool:
        return bool(self.data["links"].get(logical_id))

    def set_link(self, logical_id: str) -> None:
        self.data["links"][logical_id] = True
        self.save()


def ensure_bullets(text: str) -> str:
    lines = [ln.strip() for ln in sanitize_text(text, multiline=True).split("\n")]
    lines = [ln for ln in lines if ln]
    if len(lines) <= 1:
        return "\n".join(lines)
    if all(ln.startswith(("* ", "- ")) for ln in lines):
        return "\n".join(lines)
    return "\n".join([f"* {ln}" for ln in lines])


def plan_counts(feeds_count: int, scenarios_count: int) -> Dict[str, int]:
    return {
        "features": (1 if feeds_count > 0 else 0) + scenarios_count,
        "stories": feeds_count + scenarios_count,
        "subtasks": feeds_count * 9 + scenarios_count * 15,
    }


def main() -> int:
    args = parse_args()
    metadata = load_metadata(args.metadata)
    all_rows = read_template_rows(args.template, args.sheet)
    expected_total = args.feed_rows + args.scenario_rows
    if len(all_rows) < expected_total:
        raise ValueError(f"Template rows too short. Required at least {expected_total}, found {len(all_rows)}")

    feed_section = section_from_rows(all_rows[: args.feed_rows])
    scenario_section = section_from_rows(all_rows[args.feed_rows : args.feed_rows + args.scenario_rows])

    feeds = metadata.get("feeds", [])
    scenarios = metadata.get("scenarios", [])
    surveillance_name = metadata["surveillanceName"]
    project_key = metadata["projectKey"]
    epic_key = metadata["epicKeyForFeatures"]
    labels = metadata.get("labels", [])
    link_type = metadata.get("linkType", "Relates")

    counts = plan_counts(len(feeds), len(scenarios))
    print("Planned counts:", json.dumps(counts, indent=2))

    state = StateStore(args.state_file)

    if args.dry_run:
        print("Dry-run mode enabled. No Jira items will be created.")
        return 0

    jira_url, jira_email, jira_token, jira_auth_mode, jira_api_version = resolve_jira_credentials(args, metadata)
    if not jira_url or not jira_token:
        raise ValueError(
            "Missing Jira credentials. Provide Jira URL and token via CLI args, env vars, or metadata.jira."
        )
    if jira_auth_mode in ("basic", "auto") and not jira_email:
        raise ValueError("Basic auth mode requires jira email/username.")
    if args.auth_debug:
        print(
            "Auth config:",
            json.dumps(
                {
                    "jiraUrl": jira_url,
                    "authMode": jira_auth_mode,
                    "apiVersion": jira_api_version,
                    "emailProvided": bool(jira_email),
                    "tokenLength": len(jira_token),
                },
                indent=2,
            ),
        )

    client = JiraClient(
        base_url=jira_url,
        email=jira_email,
        token=jira_token,
        auth_mode=jira_auth_mode,
        api_version=jira_api_version,
        max_retries=args.max_retries,
        backoff=args.retry_backoff_seconds,
    )

    created_summary: Dict[str, List[str]] = {"features": [], "stories": [], "subtasks": []}

    # Feed migration: one feature, one story per feed, nine sub-tasks per story
    feed_feature_key = ""
    if feeds:
        feed_feature_id = "feed.feature"
        feed_feature_key = state.get_issue(feed_feature_id) or ""
        if not feed_feature_key:
            feed_feature_key = client.create_issue(
                project_key=project_key,
                issue_type="New Feature",
                summary=apply_placeholders(feed_section.feature, surveillance_name=surveillance_name),
                description=ensure_bullets(
                    apply_placeholders(feed_section.feature_description, surveillance_name=surveillance_name)
                ),
                epic_key=epic_key,
                labels=labels,
            )
            state.set_issue(feed_feature_id, feed_feature_key)
        if feed_section.feature_ac:
            client.set_acceptance_criteria(
                feed_feature_key,
                apply_placeholders(feed_section.feature_ac, surveillance_name=surveillance_name),
            )
        created_summary["features"].append(feed_feature_key)

        for idx, feed in enumerate(feeds, start=1):
            feed_name = sanitize_text(feed.get("name") or feed.get("feedName") or f"Feed-{idx}", multiline=False)
            story_id = f"feed.story.{feed_name}"
            story_key = state.get_issue(story_id) or ""
            if not story_key:
                story_summary = apply_placeholders(feed_section.story, surveillance_name=surveillance_name, feed_name=feed_name)
                if "FEED_NAME" not in feed_section.story and len(feeds) > 1:
                    story_summary = f"{story_summary} - {feed_name}"
                story_key = client.create_issue(
                    project_key=project_key,
                    issue_type="Story",
                    summary=story_summary,
                    description=ensure_bullets(
                        apply_placeholders(
                            feed_section.story_description,
                            surveillance_name=surveillance_name,
                            feed_name=feed_name,
                        )
                    ),
                    labels=labels,
                )
                state.set_issue(story_id, story_key)
            created_summary["stories"].append(story_key)

            link_id = f"link.story_feature.{story_key}.{feed_feature_key}"
            if not state.has_link(link_id):
                client.link_issues(story_key, feed_feature_key, link_type)
                state.set_link(link_id)

            for sub_idx, (sub_name, sub_desc, sub_ac) in enumerate(feed_section.subtasks, start=1):
                sub_id = f"feed.subtask.{feed_name}.{sub_idx}"
                sub_key = state.get_issue(sub_id) or ""
                if not sub_key:
                    sub_key = client.create_issue(
                        project_key=project_key,
                        issue_type="Sub-task",
                        summary=apply_placeholders(
                            sub_name,
                            surveillance_name=surveillance_name,
                            feed_name=feed_name,
                        ),
                        description=ensure_bullets(
                            apply_placeholders(
                                sub_desc,
                                surveillance_name=surveillance_name,
                                feed_name=feed_name,
                            )
                        ),
                        parent_key=story_key,
                        labels=labels,
                    )
                    state.set_issue(sub_id, sub_key)
                if sub_ac:
                    client.set_acceptance_criteria(
                        sub_key,
                        apply_placeholders(
                            sub_ac,
                            surveillance_name=surveillance_name,
                            feed_name=feed_name,
                        ),
                    )
                created_summary["subtasks"].append(sub_key)

    # Scenario migration: one feature, one story, fifteen sub-tasks per scenario
    for idx, scenario in enumerate(scenarios, start=1):
        scenario_name = sanitize_text(
            scenario.get("name") or scenario.get("scenarioName") or f"Scenario-{idx}",
            multiline=False,
        )
        feature_id = f"scenario.feature.{scenario_name}"
        feature_key = state.get_issue(feature_id) or ""
        if not feature_key:
            feature_key = client.create_issue(
                project_key=project_key,
                issue_type="New Feature",
                summary=apply_placeholders(
                    scenario_section.feature,
                    surveillance_name=surveillance_name,
                    scenario_name=scenario_name,
                ),
                description=ensure_bullets(
                    apply_placeholders(
                        scenario_section.feature_description,
                        surveillance_name=surveillance_name,
                        scenario_name=scenario_name,
                    )
                ),
                epic_key=epic_key,
                labels=labels,
            )
            state.set_issue(feature_id, feature_key)
        if scenario_section.feature_ac:
            client.set_acceptance_criteria(
                feature_key,
                apply_placeholders(
                    scenario_section.feature_ac,
                    surveillance_name=surveillance_name,
                    scenario_name=scenario_name,
                ),
            )
        created_summary["features"].append(feature_key)

        story_id = f"scenario.story.{scenario_name}"
        story_key = state.get_issue(story_id) or ""
        if not story_key:
            story_key = client.create_issue(
                project_key=project_key,
                issue_type="Story",
                summary=apply_placeholders(
                    scenario_section.story,
                    surveillance_name=surveillance_name,
                    scenario_name=scenario_name,
                ),
                description=ensure_bullets(
                    apply_placeholders(
                        scenario_section.story_description,
                        surveillance_name=surveillance_name,
                        scenario_name=scenario_name,
                    )
                ),
                labels=labels,
            )
            state.set_issue(story_id, story_key)
        created_summary["stories"].append(story_key)

        link_id = f"link.story_feature.{story_key}.{feature_key}"
        if not state.has_link(link_id):
            client.link_issues(story_key, feature_key, link_type)
            state.set_link(link_id)

        for sub_idx, (sub_name, sub_desc, sub_ac) in enumerate(scenario_section.subtasks, start=1):
            sub_id = f"scenario.subtask.{scenario_name}.{sub_idx}"
            sub_key = state.get_issue(sub_id) or ""
            if not sub_key:
                sub_key = client.create_issue(
                    project_key=project_key,
                    issue_type="Sub-task",
                    summary=apply_placeholders(
                        sub_name,
                        surveillance_name=surveillance_name,
                        scenario_name=scenario_name,
                    ),
                    description=ensure_bullets(
                        apply_placeholders(
                            sub_desc,
                            surveillance_name=surveillance_name,
                            scenario_name=scenario_name,
                        )
                    ),
                    parent_key=story_key,
                    labels=labels,
                )
                state.set_issue(sub_id, sub_key)
            if sub_ac:
                client.set_acceptance_criteria(
                    sub_key,
                    apply_placeholders(
                        sub_ac,
                        surveillance_name=surveillance_name,
                        scenario_name=scenario_name,
                    ),
                )
            created_summary["subtasks"].append(sub_key)

    print("Creation summary:")
    print(json.dumps(created_summary, indent=2))
    print("State file:", args.state_file)
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as err:  # pylint: disable=broad-exception-caught
        print(f"ERROR: {err}", file=sys.stderr)
        raise SystemExit(1)
